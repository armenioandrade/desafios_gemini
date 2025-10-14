import sys
modulos_path = r'C:\ARCA'
sys.path.append(modulos_path)

import os
import time
from datetime import datetime
from pathlib import Path
import pandas as pd
import pyautogui
from openpyxl import load_workbook
from selenium.webdriver.common.by import By

# Inicializar a lista para armazenar os resultados
results = []

def consulta_irregularidades(browser, download_dir):
    irregularidade = ""
    lista_ie = ler_xlsx_exportar_lista(r'C:\ARCA\Base\Empresas.xlsx', 'IE', '-')

    for ie in lista_ie:
        # Selecionar o tipo de documento
        clicar_name_selectbox(browser, 'TpDocumentoIdentificacaoCertidaoRegularidade', "INSCRIÇÃO ESTADUAL")
        time.sleep(0.5)

        clicar_name(browser, 'NuDocumentoIdentificacaoCertidaoRegularidade', 'campo_texto').send_keys(ie)

        # Clicar no botão "Localizar"
        clicar_id(browser, 'btt_localizar')
        time.sleep(2)

        # Verificar se o span "Nenhum registro encontrado" está presente
        span_elements = browser.find_elements(
            By.XPATH, "//span[text()='Nenhum registro encontrado']"
        )
        
        if span_elements:
            irregularidade = 'Nenhum registro encontrado'
        else:
            irregularidade = clicar_xpath(browser, '//*[@id="table_tabeladados"]/tbody/tr[2]/td[6]').text


        # Capturar os valores dos campos
        ie_value = browser.find_element(
            By.ID, "NuDocumentoIdentificacaoRequerente"
        ).get_attribute("value")
        razao_social_value = browser.find_element(
            By.ID, "NmRazaoSocialRequerente"
        ).get_attribute("value")

        apelido = obter_apelido(ie)

        # Adicionar os dados à lista de resultados
        results.append(
            {"CNPJ": ie_value, "RAZÃO SOCIAL": razao_social_value, "IRREGULARIDADE": irregularidade, 'COD-APELIDO': apelido}
        )
        print('linha adicionada', results[-1])
        # Adicionar um tempo de espera antes da próxima iteração, se necessário
        time.sleep(0.2)

    # Converter a lista de resultados em um DataFrame
    results_df = pd.DataFrame(results)

    # Escrever os dados em um arquivo XLSX
    file_path = rf"{download_dir}\irregularidades_contribuintes.xlsx"
    results_df.to_excel(rf"{download_dir}\irregularidades_contribuintes.xlsx", index=False)

    # Adicionar a data e hora atual na primeira linha e coluna
    data_hora_atual = datetime.now().strftime("DATA EXECUCAO %d/%m/%Y %H:%M:%S")
    workbook = load_workbook(file_path)
    sheet = workbook.active
    sheet.insert_rows(1)
    sheet.cell(row=1, column=1).value = data_hora_atual
    workbook.save(file_path)

    print("Dados salvos no arquivo irregularidades_contribuintes.xlsx")

def main():
    download_dir = criar_pasta(rf"IRREGULARIDADES\MANUAL_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}")
    browser = iniciar_browser(download_dir)
    url = chamar_url(
        browser, "https://efisco.sefaz.pe.gov.br/sfi_com_sca/PRMontarMenuAcesso"
    )
    autenticar_certificado(url)
    href_irregularidades = browser.find_element(By.LINK_TEXT, "Irregularidades do Contribuinte")
    href_irregularidades.click()
    time.sleep(2)
    consulta_irregularidades(browser, download_dir)
    time.sleep(2)
    browser.close()
    print('SEM DOWNLODAS A FAZER, AUTOMACAO ENCERRADA')

if __name__ == "__main__":
    print('Consulta Irregularidades')
    main()
